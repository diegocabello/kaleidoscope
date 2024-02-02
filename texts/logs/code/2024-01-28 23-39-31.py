import re
import os
import subprocess
import shutil
import openpyxl
from datetime import datetime
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from bigstringer import bigstringer
from text_parser import sentencer, entitier

initial_quote_color = "DFBB4C"        # light yellow
alternate_quote_color = "FFE699"      # darker yellow
initial_paragraph_color = "6691BA"    # light blue
alternate_paragraph_color = "9BC2E6"  # darker blue

def find_speaker(sentence): # is not currently in use but might use it to automatically tell which speaker is speaking
    speaker = ''
    if (0 < sentence.find('”') + 1 < len(sentence)) or (sentence[1:].find('“') >= 1):
        lister = list(x.strip(', ') for x in re.split(r'[“”]', sentence) if x)
        for index, value in enumerate(lister):
            if index == 0 and not sentence.startswith('“'):
                speaker = str(entitier(inputer = value)[0]) if entitier(inputer = value) else None
            elif index == 1 and sentence.startswith('“'):
                speaker = str(entitier(inputer = value)[0]) if entitier(inputer = value) else None
    return speaker if speaker != '' else None

paragraph_color = initial_paragraph_color     # redeclare each time
quote_color = initial_quote_color
type = 'paragraph'

paragraph_ender_pattern = '[).?!]\r\n\r\n$' # choose between `)`, `.`, `?`, `!`; optionally (\r\n\r\n), `$` adds to the end
def quote_parser(sentence):

    global type, paragraph_color, quote_color

    display_type = None
    if '“' in sentence:   #     if sentence.startswith('“'): # this is for the context `Alice said, “something.”`
        type = 'quote'
        if sentence.endswith('.\r\n\r\n'): # this is for the context `"what?" said he.`
            display_type = 'starts quote, ends paragraph'
    if type == 'quote':
        return_color = quote_color
    elif type == 'paragraph':
        return_color = paragraph_color
    if sentence.endswith('”\r\n\r\n'): # or sentence.endswith('”'):
        display_type = 'quote flipped'
        type = 'paragraph'
        quote_color = alternate_quote_color if quote_color == initial_quote_color else initial_quote_color
        paragraph_color = initial_paragraph_color
    elif re.search(paragraph_ender_pattern, sentence):
        display_type = 'paragraph flipped'
        paragraph_color = alternate_paragraph_color if paragraph_color == initial_paragraph_color else initial_paragraph_color
        quote_color = initial_quote_color

    return {'color': return_color, 'type': type if not display_type else display_type}

def main(book_name=None):

    global type, paragraph_color, quote_color

    def solid_fill(color):
        return PatternFill(start_color=color, end_color=color, fill_type='solid')

    if not book_name:
        book_name = 'pride and prejudice'
    workbook_path = f'c:\\users\\diego\\documents\\my stuff\\programming stuff\\babel\\texts\\{book_name}.xlsx'
    run_log_time = datetime.now().strftime("%Y-%m-%d %H-%M-%S")

    columns = [
        ('bit', 'information'),
        ('sentence', 'information'),
        ('type', 'information'),
        ('character length', 'information'),
        ('setting', 'until changed'),
        ('style', 'until changed'),
        ('frame', 'one shot'),
        ('speaker', 'one shot'),
        ('display', 'one shot')
    ]

    header_border = Border(left=Side(style='thin', color='575757'), right=Side(style='thin', color='575757'), top=Side(style='thin', color='575757'), bottom=Side(style='medium'))
    border = Border(left=Side(style='thin', color='575757'), right=Side(style='thin', color='575757'), top=Side(style='thin', color='575757'), bottom=Side(style='thin', color='575757'))

    # IMPORTING AND PARSING
    text = bigstringer(f'c:\\users\\diego\\documents\\my stuff\\programming stuff\\babel\\texts\\{book_name}.txt')
    chapter_list = ['CHAPTER ' + chapter for chapter in text.split('CHAPTER') if chapter][0:1]
    #chapter_list = [chapter[(chapter.find('. \r\n')+1):] for chapter in chapter_list]
    chapter_list = [chapter.split('”\r\n\r\n“') for chapter in chapter_list] # these are not regular quote marks, they are fancy unicode quotes; chapter.split('”\r\n\r\n“')
    
    if os.path.exists(workbook_path):
        subprocess.run(f'del "texts\\{book_name}.xlsx"', shell=True)
    workbook = openpyxl.Workbook()

    for count, chapter in enumerate(chapter_list):
        sheet_name = str(count + 1)

        paragraph_color = initial_paragraph_color     # redeclare the colors to reset it for each chapter
        quote_color = initial_quote_color
        type = 'paragraph'
        cut_by_value = 40

        # PARSING 
        # basically spacy doesnt support open or closed double quotes and it also sucks so i have to parse it 
        len_ch_min_one = (len(chapter) - 1)
        for counter, quot in enumerate(chapter):                # readding quote marks 
            if counter != 0 or len_ch_min_one:              # this is first because it causes the code to check itself the least amount of times
                chapter[counter] = '“' + quot + '”\r\n\r\n'
            elif counter == 0:                             
                chapter[counter] = quot + '”\r\n\r\n'
            elif counter == len_ch_min_one:
                chapter[counter] = '“' + quot        

        sentence_list = []
        for quot in chapter:                                    # splits the quotes into sentences
            sentence_list.extend(sentencer(inputer = quot))     
        for counter, sentence in enumerate(sentence_list):      # it has to be multiple 'for' loops each time for reassignment
            sentence_list[counter] = str(sentence)              # turns them into strings
        for counter, sentence in enumerate(sentence_list):      # replaces floating closed quotes 
            if sentence_list[counter].startswith('”'):          
                sentence_list[counter - 1] = sentence_list[counter - 1] + '”'
                sentence_list[counter] = sentence_list[counter].strip('”')
            if sentence_list[counter].isspace():   # == '\r\n':
                del sentence_list[counter]

        ques_exclam_close_qt_pattern = '[?!]”\s?$'              # either ? or ! [?!], optional space \s?, anchors pattern to end of string $
        for counter, sentence in enumerate(sentence_list):      # it could split a partial quote that ends in '!”' or '?”' as '!' and '”' or '?' and '”' so needs multiple 'for' loops
            if re.search(ques_exclam_close_qt_pattern, sentence_list[counter]): # makes sure sentences with partial quotes that end in '!”' or '?”' stay intact
                if not sentence_list[counter + 1].startswith("“"):
                    if sentence_list[counter + 1][0].islower(): # this isnt completely failproof in case the sentence goes “what on earth?” Alice screamed, or dialogue tag after
                        sentence_list[counter] = sentence_list[counter] + ' ' + sentence_list[counter + 1]                
                        del sentence_list[counter + 1]
                        print(f'\tconcatenated sentence at: sheet {count + 1}, line {counter}')
                    else:
                        print(f'\tpotential dialouge tag after which may need manual concatenation at: sheet {count + 1}, line {counter}, regex')
                        print('\t' + sentence_list[counter] + '\n\t' + sentence_list[counter + 1])
        
        # this list concatenation splits each 'sentence' into two or more if there are return newline patterns in the middle of it 
        sentence_list = [sub_sentence + '\r\n\r\n' if len(sentence.split('\r\n\r\n')) > 1 and sub_sentence != sentence.split('\r\n\r\n')[-1] else sub_sentence for sentence in sentence_list for sub_sentence in sentence.split('\r\n\r\n') if sub_sentence] 
        # and this one splits by semicolon
        sentence_list = [sub_sentence + ';' if len(sentence.split(';')) > 1 and sub_sentence != sentence.split(';')[-1] else sub_sentence 
                         for sentence in sentence_list for sub_sentence in sentence.split(';') if sub_sentence and len(sentence) > 150] 

        # WRITE TO EXCEL
        sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name) # make the sheet
        sheet.column_dimensions['B'].width = 125
        sheet.freeze_panes = 'A3'
        for column_number, value in enumerate(columns, start=1): # this is the header
            cell = sheet.cell(row=1, column=column_number)
            cell.value = value[0]
            cell.alignment, cell.font, cell.border, cell.fill  = Alignment(horizontal='center'), Font(bold=True), border, solid_fill('BFBFBF')
        for column_number, value in enumerate(columns, start=1): # this denotes the changer
            cell = sheet.cell(row=2, column=column_number)
            cell.value = value[1]
            cell.alignment, cell.border, cell.fill = Alignment(horizontal='center'), header_border, solid_fill('AD7070' if value[1] == 'one shot' else '974B4B')
        
        for counter, sentence in enumerate(sentence_list):
            if not sentence or not sentence.isspace():

                qps = quote_parser(sentence)
                color = solid_fill(qps['color'])

                columns_values = {
                    'bit': {'value': None, 'color': color},
                    'sentence': {'value': sentence, 'color': color},
                    'type': {'value': qps['type'], 'color': color},
                    'character length': {'value': len(sentence), 'color': 'c00000' if len(sentence) < cut_by_value else color},
                    'setting': {'value': None, 'color': color},
                    'style': {'value': None, 'color': color},
                    'frame': {'value': None, 'color': color},
                    'speaker': {'value': find_speaker(sentence), 'color': color},
                    'display': {'value': None, 'color': color}
                }

                row_values = []
                for key in columns_values: 
                    row_values.append(columns_values[f'{key}']['value'])
                sheet.append(row_values)

                for counter, cell in enumerate(sheet[counter + 3]):
                    cell.fill, cell.border = columns_values[columns[counter][0]]['color'], border # sorry that is a mess 

        workbook.save(workbook_path)
        print(f'did chapter {str(count + 1)} ')

    for sheet in ['Sheet']:
        if sheet in workbook:
            del workbook[sheet]
            workbook.save(workbook_path)

    shutil.copy(r'c:\users\diego\documents\my stuff\programming stuff\babel\super parser.py', f'c:\\users\\diego\\documents\\my stuff\\programming stuff\\babel\\texts\\logs\\code\\{run_log_time}.py')
    workbook.save(f'c:\\users\\diego\\documents\\my stuff\\programming stuff\\babel\\texts\\logs\\excel\\{book_name} {run_log_time}.xlsx')

    print('done! ')

    if os.name == 'posix':  # Unix/Linux/MacOS
        command = f'open {workbook_path}'
    elif os.name == 'nt':   # Windows
        command = f'start excel "{workbook_path}"'
    else:
        raise OSError('Unsupported operating system')

    subprocess.run(command, shell=True)
    print('launching excel... ')

if __name__ == "__main__":
    main()

