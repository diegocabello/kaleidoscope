import re
import os
import subprocess
import shutil
import openpyxl
from datetime import datetime
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from bigstringer import bigstringer
from text_parser import sentencer, entitier

initial_quote_color = "DFBB4C"        # light yellow
alternate_quote_color = "FFE699"      # darker yellow
initial_paragraph_color = "6691BA"    # light blue
alternate_paragraph_color = "9BC2E6"  # darker blue
initial_bit_color = "7A808C"          # light grey
alternate_bit_color = "1E1F21"        # dark grey 
ignore_color = "2A4225"               # green-grey  

def flatten_list(lst):
    result = []
    for item in lst:
        if isinstance(item, list):
            result.extend(flatten_list(item))
        else:
            result.append(item)
    return result

def get_speaker(sentence): # automatically tells which speaker is speaking
    speaker = ''
    if (0 < sentence.find('”') + 1 < len(sentence)) or (sentence[1:].find('“') >= 1):
        lister = list(x.strip(', ') for x in re.split(r'[“”]', sentence) if x)
        for index, value in enumerate(lister):
            if index == 0 and not sentence.startswith('“'):
                try:
                    prefix = re.compile(r'(Mr\.|Mrs\.|Ms\.|Miss)\s').search(value).group(1)
                except AttributeError:
                    prefix = None
                speaker = prefix if prefix else '' + str(entitier(inputer = value)[0]) if entitier(inputer = value) else None
            elif index == 1 and sentence.startswith('“'):
                speaker = str(entitier(inputer = value)[0]) if entitier(inputer = value) else None
    return speaker if speaker != '' else None

paragraph_color = initial_paragraph_color     # redeclare each time
quote_color = initial_quote_color
type = 'paragraph'

paragraph_ender_pattern = '[).?!]\r\n\r\n$' # choose between `)`, `.`, `?`, `!`; optionally (\r\n\r\n), `$` adds to the end | need to add `—` as an 'or' option 
def section_parser(sentence):

    global type, paragraph_color, quote_color

    if '“' in sentence:       # if sentence.startswith('“'): # this is for the context `Alice said, “something.”`
        type = 'quote'
    else:
        type = 'paragraph'    # if sentence.endswith('.\r\n\r\n'): # this is for the context `"what?" said he.`
    if type == 'quote':
        return_color = quote_color
        quote_color = alternate_quote_color if quote_color == initial_quote_color else initial_quote_color
        paragraph_color = initial_paragraph_color
    elif type == 'paragraph':
        return_color = paragraph_color
        paragraph_color = alternate_paragraph_color if paragraph_color == initial_paragraph_color else initial_paragraph_color
        quote_color = initial_quote_color

    return {'color': return_color, 'type': type}

def parse_book(book_name=None):

    global type, paragraph_color, quote_color

    def solid_fill(color):
        return PatternFill(start_color=color, end_color=color, fill_type='solid')
    header_border = Border(left=Side(style='thin', color='575757'), right=Side(style='thin', color='575757'), top=Side(style='thin', color='575757'), bottom=Side(style='medium'))
    border = Border(left=Side(style='thin', color='575757'), right=Side(style='thin', color='575757'), top=Side(style='thin', color='575757'), bottom=Side(style='thin', color='575757'))

    if not book_name:
        book_name = 'pride and prejudice'
    workbook_path = f'texts\\{book_name}.xlsx'
    run_log_time = datetime.now().strftime("%Y-%m-%d %H-%M-%S")

    columns = [
        ('bit', 'information'),
        ('sentence', 'information'),
        ('type', 'information'),
        ('character length', 'information'),
        ('setting', 'until changed'),
        ('temp setting', 'one shot'),
        ('style', 'until changed'),
        ('display', 'one shot'), 
        ('frame', 'one shot'),
        ('speaker', 'one shot'),
        ('spoken to', 'one shot'),
        ('also present', 'information'), 
        ('remove presence', 'information')
    ]

    index_of_also_present = next((index for index, column in enumerate(columns) if column[0] == 'also present'), None)

    cut_by_value = 40

    shutil.copy(r'super parser.py', f'texts\\logs\\code\\{run_log_time}.py')

    # IMPORTING AND PARSING
    text = bigstringer(f'texts\\{book_name}.txt')
    chapter_list = ['CHAPTER ' + chapter for chapter in text.split('CHAPTER') if chapter][0:30]
    #chapter_list = [chapter[(chapter.find('. \r\n')+1):] for chapter in chapter_list]
    for count, chapter in enumerate(chapter_list):
        chapter_list[count] = [section + '\r\n\r\n' for section in chapter.split('\r\n\r\n') if section] # these are not regular quote marks, they are fancy unicode quotes; chapter.split('”\r\n\r\n“')
    
    if os.path.exists(workbook_path):
        subprocess.run(f'del "texts\\{book_name}.xlsx"', shell=True)
    workbook = openpyxl.Workbook()

    for count, chapter in enumerate(chapter_list):                                                              #    ====CHAPTER====

        sheet_name = str(count + 1) 
        start_index = 3

        paragraph_color, quote_color, bit_color = initial_paragraph_color, initial_quote_color, alternate_bit_color   # redeclare the colors to reset it for each chapter
        prev_type, speaker = None, None
        do_the_thing_flag = True
        type = 'paragraph'
        present_list = []

        # WRITE HEADER TO EXCEL
        sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name) # make the sheet
        sheet.column_dimensions['B'].width = 125
        sheet.freeze_panes = 'A3'
        for column_number, value in enumerate(columns, start=1): # this is the header
            cell = sheet.cell(row=1, column=column_number)
            cell.value = value[0]
            cell.alignment, cell.font, cell.border, cell.fill  = Alignment(horizontal='center'), Font(bold=True), border, solid_fill('BFBFBF')
        for column_number, value in enumerate(columns, start=1): # this denotes the changer
            cell = sheet.cell(row=2, column=column_number)
            cell.value = value[1]
            cell.alignment, cell.border, cell.fill = Alignment(horizontal='center'), header_border, solid_fill('AD7070' if value[1] == 'one shot' else '974B4B')

        # PARSE SECTIONS
        for section in chapter:                                                                                 #    ====SECTION====
            sp = section_parser(section)

            prev_type = typer if prev_type else None    # pushes it back if its there

            color = solid_fill(sp['color'])
            typer = sp['type']

            if not prev_type:                           # defines it if its not there 
                prev_type = typer

            if type == 'paragraph':
                do_the_thing_flag = True

            spoken_to = None
            if typer == 'quote':
                if prev_type == 'quote':
                    if do_the_thing_flag:
                        speaker1 = speaker
                        speaker2 = get_speaker(section) if get_speaker(section) else None  # pushes it back and redefines it if there
                        do_the_thing_flag = False
                        speaker_list = [speaker2, speaker1]
                    speaker_list.reverse()
                    speaker = speaker_list[1]
                    spoken_to = speaker_list[0]
                if get_speaker(section):
                    speaker = get_speaker(section)
                    present_list.append(speaker)
            else:
                speaker = None

            # PARSE SENTENCES
            sentence_list = [str(sentence) for sentence in sentencer(inputer=section)]                          #    ====SENTENCE====

            for counter, sentence in enumerate(sentence_list):      # spacy doesnt correctly parse open or closed double quotes sometimes
                if sentence_list[counter].startswith('”'):          # replaces floating closed quotes  
                    sentence_list[counter - 1] = sentence_list[counter - 1] +  '”'
                    sentence_list[counter] = sentence_list[counter].strip('”')
                if sentence_list[counter].isspace():   # == '\r\n':
                    del sentence_list[counter]

            ques_exclam_close_qt_pattern = '[?!]”\s?$'              # either ? or ! [?!], optional space \s?, anchors pattern to end of string $
            for counter, sentence in enumerate(sentence_list):      # it could split a partial quote that ends in '!”' or '?”' as '!' and '”' or '?' and '”' so needs multiple 'for' loops
                if re.search(ques_exclam_close_qt_pattern, sentence_list[counter]): # makes sure sentences with partial quotes that end in '!”' or '?”' stay intact
                    if not sentence_list[counter + 1].startswith("“") and sentence_list[counter + 1][0].islower(): # this isnt completely failproof in case the sentence goes “what on earth?” Alice screamed, or dialogue tag after
                        sentence_list[counter] += ' ' + sentence_list.pop(counter + 1)                        
                        print(f'\tconcatenated sentence at: sheet {count + 1}, line {start_index}')
                    else:
                        print(f'\tpotential dialouge tag after which may need manual concatenation at: sheet {count + 1}, line {start_index}, regex')
                        print('\t' + sentence_list[counter] + '\n\t' + sentence_list[counter + 1])
            
            # this splits by semicolon if it is longer then a certain length 
            def split_and_tack_on_properly(inputer = '', splitter = ''):
                splitted = inputer.split(splitter)
                return [sub_sentence + splitter if sub_sentence != splitted[-1] else sub_sentence for sub_sentence in splitted if sub_sentence] 

            return_list = []  
            for sentence in sentence_list:
                to_extend_by = sentence
                if len(sentence) > 150 and sentence.find(';') not in [0, -1, (len(sentence) - 1)]:
                    to_extend_by = split_and_tack_on_properly(inputer = sentence, splitter = ';')
                return_list.extend(to_extend_by)
            sentence_list = return_list

            # WRITE SENTENCES TO EXCEL 
            bit_color = alternate_bit_color if bit_color == initial_bit_color else initial_bit_color
            bit_counter = 0

            for sentence_counter, sentence in enumerate(sentence_list):
                if not sentence or not sentence.isspace():

                    # speaker, color, and typer are defined above; if it is None it has to be defined in an if/else block 

                    paragraph_blackout = color
                    if typer == 'quote':
                        if sentence_counter == 0:
                            frame = 'speaker' 
                        else:
                            frame, speaker, spoken_to = None, None, None
                            paragraph_blackout = solid_fill(ignore_color) # 8C887A
                    elif typer == 'paragraph': 
                        paragraph_blackout = solid_fill(ignore_color) # 8C887A
                        frame = None
                    else:
                        frame = None

                    general_blackout, present_blackout = color, solid_fill('d8d8d8')
                    if bit_counter > 0:
                        general_blackout, present_blackout = solid_fill(ignore_color), solid_fill(ignore_color)

                    columns_values = {
                        'bit': {'value': None, 'color': solid_fill(bit_color)},
                        'sentence': {'value': sentence + ' ', 'color': color},
                        'type': {'value': typer, 'color': color},
                        'character length': {'value': len(sentence), 'color': solid_fill('c00000') if len(sentence) < cut_by_value else color},
                        'setting': {'value': None, 'color': general_blackout},
                        'temp setting': {'value': None, 'color': general_blackout}, 
                        'style': {'value': None, 'color': general_blackout},
                        'display': {'value': None, 'color': general_blackout}, 
                        'frame': {'value': frame, 'color': paragraph_blackout},
                        'speaker': {'value': speaker, 'color': paragraph_blackout},
                        'spoken to': {'value': spoken_to, 'color': paragraph_blackout},
                        'also present': {'value': None, 'color': present_blackout}, 
                        'remove presence' : {'value': None, 'color': present_blackout}
                    }

                    bit_counter = bit_counter + len(sentence)
                    if bit_counter > 150:
                        bit_counter = 0
                        if sentence_counter != (len(sentence_list) - 1):
                            bit_color = alternate_bit_color if bit_color == initial_bit_color else initial_bit_color

                    row_values = []
                    for key in columns_values: 
                        row_values.append(columns_values[f'{key}']['value'])
                    sheet.append(row_values)

                    for cell_counter, cell in enumerate(sheet[start_index]):
                        cell.fill, cell.border = columns_values[columns[cell_counter][0]]['color'], border 

                start_index = start_index + 1 

        workbook[3][index_of_also_present] = ', '.join(present_list)

        workbook.save(workbook_path)
        print(f'did chapter {str(count + 1)} ')

    for sheet in ['Sheet']:
        if sheet in workbook:
            del workbook[sheet]
            workbook.save(workbook_path)

    workbook.save(f'texts\\logs\\excel\\{book_name} {run_log_time}.xlsx')

    print('done! ')

    if os.name == 'posix':  # Unix/Linux/MacOS
        command = f'open {workbook_path}'
    elif os.name == 'nt':   # Windows
        command = f'start excel "{workbook_path}"'
    else:
        raise OSError('Unsupported operating system')

    subprocess.run(command, shell=True)
    print('launching excel... ')

if __name__ == "__main__":
    parse_book()


