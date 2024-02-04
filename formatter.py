import subprocess

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

initial_bit_color = "7A808C"          # light grey
ignore_color = "2A4225"               # green-grey  

character_descriptions = {'Elizabeth': 'a young woman in upperclass fancy victorian dress',
                          'Darcy': 'a young man in victorian england',
                          'Mr. Bennet': 'a middle-aged, older man in victorian england',
                          'Mrs. Bennet': 'a middle aged, older woman in victorian england fancy dress'}

def charachter_with_description(character=''):
    return f'{character}, {character_descriptions[character]} ' if character in character_descriptions else character

def format_book(book_name=None):

    if not book_name:
        book_name = 'pride and prejudice'
    workbook = openpyxl.load_workbook(f'texts\\{book_name}.xlsx')
    workbook = workbook.worksheets[0:1]

    list_of_keys = []
    dicto_change = {}
    dicto_values = {}

    list_of_prompts = []
    list_of_bits = []

    present_list = []

    with open(f'prompts\\{book_name} prompts.txt', 'w') as filer:
        filer.write('')

    initialize_flag = True
    for sheet_counter, sheet in enumerate(workbook): 
        if initialize_flag:
            for col, cell in enumerate(sheet[1], start=1):
                dicto_change[f'{str(cell.value)}'] = sheet.cell(row=2, column=col).value
                dicto_values[f'{str(cell.value)}'] = None
                list_of_keys.append(str(cell.value))
            for col, cell in enumerate(sheet[2], start=1): 
                dicto_change[list_of_keys[col - 1]] = str(cell.value) if cell.value != None else None
            print(dicto_change)
            print(list_of_keys)
            initialize_flag = False
        
        bit_counter = 1
        bit_color = initial_bit_color
        bit_values = {key: None for key in dicto_values} 
        return_string = ''

        for row in sheet.iter_rows(min_row = 3):

            for cell_counter, cell in enumerate(row):
                
                if cell_counter == 0:

                    prev_bit_color = bit_color 
                    bit_color = cell.fill.start_color.rgb[2:]

                    if prev_bit_color != bit_color:                     # at the change of the bit 

                        if bit_values['frame'] == 'speaker':            # formats

                            also_present_string = ''
                            if len(present_list) != 0:
                                for character in present_list:
                                    also_present_string = also_present_string + charachter_with_description(character)

                            return_string = return_string + f'\
{charachter_with_description(bit_values["speaker"])} is speaking to {charachter_with_description(bit_values["spoken to"])} \
in {bit_values["setting"]} \
{"also present in the background are" + also_present_string if also_present_string != "" else ""} \
style: {bit_values["style"]} \
--ar 16:9 '

                        else:
                            for value in bit_values:
                                if bit_values[value] != None:
                                    return_string = return_string + f'{value}: {bit_values[value]} \n'
                            return_string = return_string + ' --ar 16:9 '

                        list_of_prompts.append(return_string)           # writes
                        list_of_bits.append(bit_values['sentence'])
                        with open(f'prompts\\{book_name} prompts.txt', 'a') as filer:
                            filer.write(f'**chapter {sheet_counter + 1} bit {bit_counter}**\n'+ return_string + '\n\n')

                        return_string = f''                             # resets 
                        bit_values = {key: None for key in dicto_values} 
                        bit_counter += 1


                elif cell.fill.start_color.rgb[2:] != ignore_color:

                    cell_value = str(cell.value).strip() if cell.value != None else None
                    category = list_of_keys[cell_counter]

                    one_shot_bool = dicto_change[category] == 'one shot'            # isnt either/or, there are other options 
                    until_changed_bool = dicto_change[category] == 'until changed'   
                    isnt_none_bool = cell_value is not None  

                    if until_changed_bool and isnt_none_bool and category not in ['also present', 'remove presence']:            # if it exists and is marked until changed
                        bit_values[category] = cell_value
                        dicto_values[category] = cell_value
                    elif until_changed_bool and not isnt_none_bool and category not in ['also present', 'remove presence']:      # if it's none and is marked until changed 
                        bit_values[category] = dicto_values[category] 
                    elif (one_shot_bool or category == 'sentence') and isnt_none_bool:                                           # regular data mark
                        bit_values[category] = bit_values[category] + cell_value if bit_values[category] != None else cell_value
                    elif category == 'also present':
                        if cell_value:
                            present_list.extend(cell_value.split(', '))
                    elif category == 'remove presence':
                        if cell_value:
                            to_remove_list = cell_value.split(', ')
                            for character in to_remove_list: 
                                if character in present_list:
                                    present_list.remove(character)
                    

        dicto_values = {key: None for key in dicto_values}              # resets at the end of each sheet
        print(f'did sheet {str(sheet_counter + 1)} ')

    subprocess.run(f'start notepad prompts\\{book_name} prompts.txt', shell=True)

    return {'prompts': list_of_prompts, 'bits': list_of_bits}

if __name__ == "__main__":
    format_prompts()

